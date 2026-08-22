# database/gst_export.py
"""
GST Summary Excel Generator — EstateHub
========================================
Produces a structured GST summary report for one financial year.

Two sheets:
  * "Taxable Supplies" — month-by-month (GSTR-1 Table 4/5 shape)
  * "Summary" — one FY-total row: taxable value, tax collected, exempt value,
                total turnover

Data source: fn_gst_summary_fy(p_society_id, p_fy)
"""

from __future__ import annotations
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


_FONT_BODY = Font(name="Arial", size=9)
_FONT_HEADER = Font(name="Arial", size=9, bold=True)
_FONT_TITLE = Font(name="Arial", size=10, bold=True)
_FONT_TOTAL = Font(name="Arial", size=9, bold=True)

_FILL_HEADER = PatternFill("solid", fgColor="D9E1F2")
_FILL_TOTAL = PatternFill("solid", fgColor="E2EFDA")

_ALIGN_C = Alignment(horizontal="center", vertical="center")
_ALIGN_L = Alignment(horizontal="left", vertical="center")
_ALIGN_R = Alignment(horizontal="right", vertical="center")

_THIN = Side(style="thin")
_BORDER_ALL = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_FMT_AMT = '#,##0.00;[Red](#,##0.00);"-"'
_FMT_DATE = "MMM-YYYY"

_COL_WIDTHS = {
    "A": 14, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16, "G": 20, "H": 18,
}


def _apply_header(ws, row: int, widths: dict) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[row].height = 22


def _write_taxable_supplies_sheet(ws, rows: list[dict], fy: int) -> int:
    ws.cell(row=1, column=1, value=f"GST Summary — Taxable Supplies FY {fy}-{fy+1}")
    ws.cell(row=1, column=1).font = _FONT_TITLE

    headers = [
        "Month", "Taxable Value", "CGST Collected",
        "SGST Collected", "Total GST", "Exempt Value",
        "Bills GST Applicable", "Bills Exempt",
    ]
    for col, hdr in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=hdr)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_C
        cell.border = _BORDER_ALL

    r = 3
    t_taxable = 0.0
    t_cgst = 0.0
    t_sgst = 0.0
    t_exempt = 0.0
    t_gst_bills = 0
    t_exempt_bills = 0

    for row in rows:
        period = row.get("period_month")
        taxable = float(row.get("taxable_value", 0) or 0)
        cgst = float(row.get("cgst_collected", 0) or 0)
        sgst = float(row.get("sgst_collected", 0) or 0)
        exempt = float(row.get("exempt_value", 0) or 0)
        gst_bills = int(row.get("total_bills_gst_applicable", 0) or 0)
        exempt_bills = int(row.get("total_bills_exempt", 0) or 0)

        values = [
            period,
            taxable,
            cgst,
            sgst,
            cgst + sgst,
            exempt,
            gst_bills,
            exempt_bills,
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = _FONT_BODY
            cell.border = _BORDER_ALL
            if col == 1:
                cell.alignment = _ALIGN_C
                cell.number_format = _FMT_DATE
            elif col in (2, 3, 4, 5, 6):
                cell.alignment = _ALIGN_R
                cell.number_format = _FMT_AMT
            else:
                cell.alignment = _ALIGN_C
        r += 1

        t_taxable += taxable
        t_cgst += cgst
        t_sgst += sgst
        t_exempt += exempt
        t_gst_bills += gst_bills
        t_exempt_bills += exempt_bills

    totals = [
        "TOTAL", t_taxable, t_cgst, t_sgst, t_cgst + t_sgst,
        t_exempt, t_gst_bills, t_exempt_bills,
    ]
    for col, val in enumerate(totals, start=1):
        cell = ws.cell(row=r, column=col, value=val)
        cell.font = _FONT_TOTAL
        cell.fill = _FILL_TOTAL
        cell.border = _BORDER_ALL
        if col == 1:
            cell.alignment = _ALIGN_L
        elif col in (2, 3, 4, 5, 6):
            cell.alignment = _ALIGN_R
            cell.number_format = _FMT_AMT
        else:
            cell.alignment = _ALIGN_C
    return r


def _write_summary_sheet(ws, rows: list[dict], fy: int) -> None:
    ws.cell(row=1, column=1, value=f"GST Summary — FY {fy}-{fy+1}")
    ws.cell(row=1, column=1).font = _FONT_TITLE

    headers = [
        "Financial Year", "Taxable Value", "CGST Collected",
        "SGST Collected", "Total GST Collected", "Exempt Value",
        "Total Turnover", "GST Bills", "Exempt Bills",
    ]
    for col, hdr in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=hdr)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_C
        cell.border = _BORDER_ALL

    t_taxable = sum(float(x.get("taxable_value", 0) or 0) for x in rows)
    t_cgst = sum(float(x.get("cgst_collected", 0) or 0) for x in rows)
    t_sgst = sum(float(x.get("sgst_collected", 0) or 0) for x in rows)
    t_exempt = sum(float(x.get("exempt_value", 0) or 0) for x in rows)
    t_gst_bills = sum(int(x.get("total_bills_gst_applicable", 0) or 0) for x in rows)
    t_exempt_bills = sum(int(x.get("total_bills_exempt", 0) or 0) for x in rows)

    values = [
        f"FY {fy}-{fy+1}",
        t_taxable,
        t_cgst,
        t_sgst,
        t_cgst + t_sgst,
        t_exempt,
        t_taxable + t_cgst + t_sgst + t_exempt,
        t_gst_bills,
        t_exempt_bills,
    ]
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=3, column=col, value=val)
        cell.font = _FONT_BODY
        cell.border = _BORDER_ALL
        if col == 1:
            cell.alignment = _ALIGN_L
        elif col in (2, 3, 4, 5, 6, 7):
            cell.alignment = _ALIGN_R
            cell.number_format = _FMT_AMT
        else:
            cell.alignment = _ALIGN_C


def generate_gst_summary_excel(
    db,
    society_id: int,
    fy: int,
    filename_prefix: str = "GSTSummary",
) -> bytes:
    """Builds the GST summary workbook for one society FY."""
    from database.db_manager import db as _db
    if db is None:
        db = _db

    rows = db._execute(
        "SELECT * FROM fn_gst_summary_fy(%s,%s)",
        (society_id, fy), fetch_all=True,
    ) or []

    wb = Workbook()
    wb.remove(wb.active)

    ws_ts = wb.create_sheet(title="Taxable Supplies")
    _apply_header(ws_ts, 2, _COL_WIDTHS)
    _write_taxable_supplies_sheet(ws_ts, rows, fy)

    ws_sum = wb.create_sheet(title="Summary")
    _write_summary_sheet(ws_sum, rows, fy)

    filename = f"{filename_prefix}_FY{fy}-{fy+1}.xlsx"
    wb._gst_filename = filename  # type: ignore[attr-defined]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
