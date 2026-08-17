# database/cashbook_export.py
"""
Cashbook Excel Generator — EstateHub
======================================
Produces the traditional paired Indian society cashbook, one sheet per
month, all months of a financial year in a single workbook — matching the
CB2025-2026.xlsx reference layout supplied 2026-08.

DEPENDS ON: fn_cashbook_paired_v3 (see estatehub.sql), which requires the
entry_side migration to be deployed first. Until then this module will
raise if entry_side doesn't exist on transactions — it does NOT silently
fall back to fn_cashbook_paired_v2's buggy drcr_account-inferred join,
since that would reproduce the exact bug this rewrite exists to fix.

Column source (2026-08): every money-writing function now writes a
bank/cash-completing leg ONLY for non-cash modes (fn_resolve_bank_leg) —
a cash-mode transaction has exactly one leg, posted straight to the real
income/expense/asset account, never to CiH. fn_cashbook_paired_v3's
output columns are named cr_*/dr_* accordingly (not the old rc_*/pc_*),
and its running-balance column is `cih_running`.

Column layout (A–O) — unchanged from the original single-month version:
  A  Date (receipt side)
  B  Receipt A/c name       (accounts.tab_name)
  C  Receipt Particulars    (acc_particulars + payment_gateway_id)
  D  Receipt L.F. No.       (ledger folio = accounts.id)
  E  Receipt Cash           (mode='cash', entry_side='Cr')
  F  Receipt Chq/UPI        (mode<>'cash', entry_side='Cr', informational)
  G  Receipt Running Total  (running sum of col E — cash receipts only)
  H  Date (payment side)
  I  Payment A/c name       (accounts.tab_name)
  J  Payment Particulars    (acc_particulars + payment_gateway_id)
  K  Payment L.F. No.       (accounts.id)
  L  Payment Cash           (mode='cash', entry_side='Dr')
  M  Payment Chq            (mode<>'cash', entry_side='Dr', informational)
  N  Payment Running Total  (running sum of col L — cash payables only)
  O  Balance                (= G − N, physical cash in hand)

Row structure per month sheet:
  Row 1: blank
  Row 2: A2=filename, C2=society_name, E2='Society', F2='CASHBOOK', G2='PAN:', H2=PAN,
          J2='Asst.Yr.', K2=year_range, L2='Month', M2=month_abbrev
  Row 3: blank
  Row 4: Column headers
  Row 5: Balance B/F row — ONLY 'Balance' (col B/I) + 'B/F' (col C/J) + the
          opening amount + running-total seeds. No other data on this row.
  Row 6+: Data rows (receipt side / payment side; the side with no entry on
          a given row is left entirely blank)
  Last:  Balance C/F row — ONLY 'Balance' + 'C/F' + closing cash amount.
          No other data on this row.

FY grouping: one workbook per financial year (Apr–Mar), one sheet per
month named by month abbreviation (Apr, May, ... Mar), each month's
opening balance carried from the prior month's closing balance, the very
first month's opening balance coming from `brought_forward`.

Entity/user scoping: entity_id / entity_role passed straight through
to fn_cashbook_paired_v3, so the same generator serves both:
  - Admin portal: entity_id=None (all) or a specific user's underlying
    entity_id when 'ALL' is not selected
  - Owner/Vendor/Security portals: entity_id defaults to current user's
    linked entity_id, entity_role fixed to the portal's role

NOTE: "Intelligent grouping" (individual apartment dues -> single
'Apartment maintenance' line, individual vendor payments -> 'Vendors',
etc.) is NOT implemented in this pass. It needs a defined account-level
grouping label before it can be built without hardcoding account-name
matches, which would be fragile against the registry-driven account list.
This generator currently always shows the underlying per-account rows.
"""

from __future__ import annotations
import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet


_FONT_BODY   = Font(name="Arial", size=9)
_FONT_HEADER = Font(name="Arial", size=9, bold=True)
_FONT_TITLE  = Font(name="Arial", size=10, bold=True)
_FONT_BF_CF  = Font(name="Arial", size=9, bold=True, italic=True)

_FILL_HEADER = PatternFill("solid", fgColor="D9E1F2")
_FILL_BF     = PatternFill("solid", fgColor="E2EFDA")
_FILL_CF     = PatternFill("solid", fgColor="FCE4D6")
_FILL_ALT    = PatternFill("solid", fgColor="F7F7F7")

_ALIGN_C = Alignment(horizontal="center", vertical="center")
_ALIGN_L = Alignment(horizontal="left",   vertical="center")
_ALIGN_R = Alignment(horizontal="right",  vertical="center")

_THIN = Side(style="thin")
_BORDER_ALL = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_FMT_DATE = "DD-MMM-YY"
_FMT_AMT  = '#,##0.00;[Red](#,##0.00);"-"'

_COL_WIDTHS = {
    "A": 10, "B": 10, "C": 28, "D":  6,
    "E":  9, "F":  9, "G": 10,
    "H": 10, "I": 10, "J": 28, "K":  6,
    "L":  9, "M":  9, "N": 10, "O": 10,
}


def _particulars(row_dict: dict, prefix: str) -> str:
    p  = (row_dict.get(f"{prefix}particulars") or "").strip()
    gw = (row_dict.get(f"{prefix}cheque_no") or "").strip()
    return f"{p} [{gw}]" if gw else p


def _write_month_sheet(
    ws: Worksheet,
    rows: list[dict],
    opening_balance: float,
    society_name: str,
    pan: str,
    asst_year: str,
    month_dt: date,
    filename: str,
) -> float:
    """Writes one month's cashbook onto `ws`. Returns the closing balance."""

    for col_letter, width in _COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 6

    month_abbrev = month_dt.strftime("%b")
    title_data = {
        1: filename, 3: society_name, 5: "Society", 6: "CASHBOOK",
        7: "PAN:", 8: pan, 10: "Asst.Yr.", 11: asst_year,
        12: "Month", 13: month_abbrev,
    }
    for col, val in title_data.items():
        cell = ws.cell(row=2, column=col, value=val)
        cell.font, cell.alignment = _FONT_TITLE, _ALIGN_C

    ws.row_dimensions[3].height = 6

    headers = {
        1: "Date", 2: "Receipt A/c", 3: "Receipt Particulars", 4: "L.F.",
        5: "Cash", 6: "Chq./UPI", 7: "Total",
        8: "Date", 9: "Payment A/c", 10: "Payment Particulars", 11: "L.F.",
        12: "Cash", 13: "Chq.", 14: "Total", 15: "Balance",
    }
    for col, hdr in headers.items():
        cell = ws.cell(row=4, column=col, value=hdr)
        cell.font, cell.fill = _FONT_HEADER, _FILL_HEADER
        cell.alignment, cell.border = _ALIGN_C, _BORDER_ALL

    # Fixed (2026-08): B/F account column now shows 'CiH' — matching
    # CB2024-2025.xlsx (B6='CiH', C6='B/F') and the Cashbook's only-CiH-
    # has-a-B/F-row rule — rather than the generic label 'Balance', which
    # didn't identify which account this opening balance belongs to.
    bf_row = {
        1: month_dt, 2: "CiH", 3: "B/F",
        5: opening_balance if opening_balance >= 0 else None,
        12: abs(opening_balance) if opening_balance < 0 else None,
    }
    for col, val in bf_row.items():
        cell = ws.cell(row=5, column=col, value=val)
        cell.font, cell.fill = _FONT_BF_CF, _FILL_BF
        cell.alignment = _ALIGN_R if col in (5, 12) else _ALIGN_L
        cell.border = _BORDER_ALL
        if col == 1:
            cell.number_format = _FMT_DATE
        elif col in (5, 12):
            cell.number_format = _FMT_AMT

    for col, formula in [(7, "=E5"), (14, "=L5"), (15, "=G5-N5")]:
        cell = ws.cell(row=5, column=col, value=formula)
        cell.font, cell.fill = _FONT_BF_CF, _FILL_BF
        cell.border, cell.alignment = _BORDER_ALL, _ALIGN_R
        cell.number_format = _FMT_AMT

    current_row = 6
    for i, r in enumerate(rows):
        fill = _FILL_ALT if i % 2 == 0 else None
        row_data = {
            1: r.get("row_date") or None,
            2: r.get("cr_account_name") or None,
            3: _particulars(r, "cr_") or None,
            4: r.get("cr_acc_id") or None,
            5: float(r.get("cr_cash") or 0) or None,
            6: float(r.get("cr_chq") or 0) or None,
            8: r.get("row_date") or None,
            9: r.get("dr_account_name") or None,
            10: _particulars(r, "dr_") or None,
            11: r.get("dr_acc_id") or None,
            12: float(r.get("dr_cash") or 0) or None,
            13: float(r.get("dr_chq") or 0) or None,
        }
        for col, val in row_data.items():
            cell = ws.cell(row=current_row, column=col, value=val)
            cell.font = _FONT_BODY
            cell.fill = fill or PatternFill()
            cell.border = _BORDER_ALL
            cell.alignment = _ALIGN_R if col in (5, 6, 12, 13) else _ALIGN_L
            if col in (1, 8) and val:
                cell.number_format = _FMT_DATE
            elif col in (5, 6, 12, 13):
                cell.number_format = _FMT_AMT

        prev = current_row - 1
        for col, formula in [
            (7,  f'=G{prev}+IF(E{current_row}<>"",E{current_row},0)'),
            (14, f'=N{prev}+IF(L{current_row}<>"",L{current_row},0)'),
            (15, f"=G{current_row}-N{current_row}"),
        ]:
            cell = ws.cell(row=current_row, column=col, value=formula)
            cell.font, cell.fill = _FONT_BODY, (fill or PatternFill())
            cell.border, cell.alignment = _BORDER_ALL, _ALIGN_R
            cell.number_format = _FMT_AMT
        current_row += 1

    cf_row = current_row
    prev = cf_row - 1
    # Same 'CiH' fix as the B/F row above (was the generic label 'Balance').
    ws.cell(row=cf_row, column=9,  value="CiH")
    ws.cell(row=cf_row, column=10, value="C/F")
    ws.cell(row=cf_row, column=12, value=f"=O{prev}")
    ws.cell(row=cf_row, column=7,  value=f"=G{prev}")
    ws.cell(row=cf_row, column=14, value=f"=N{prev}+L{cf_row}")
    ws.cell(row=cf_row, column=15, value=f"=G{cf_row}-N{cf_row}")
    for col in [7, 9, 10, 12, 14, 15]:
        cell = ws.cell(row=cf_row, column=col)
        cell.font, cell.fill = _FONT_BF_CF, _FILL_CF
        cell.border = _BORDER_ALL
        cell.alignment = _ALIGN_L if col in (9, 10) else _ALIGN_R
        if col in (7, 12, 14, 15):
            cell.number_format = _FMT_AMT

    ws.freeze_panes = "A5"

    # Closing balance is computed here in Python (not read back from an
    # Excel formula), so the next month's opening balance doesn't depend on
    # spreadsheet recalculation having happened.
    closing_balance = opening_balance + sum(
        float(r.get("cr_cash") or 0) - float(r.get("dr_cash") or 0) for r in rows
    )
    return closing_balance


def generate_cashbook_excel_fy(
    db,
    society_id: int,
    fy: int,                      # e.g. 2025 for FY2025 (Apr 2025-Mar 2026)
    entity_id: int | None = None,
    entity_role: str | None = None,
    filename_prefix: str = "Cashbook",
) -> bytes:
    """
    Builds a full-FY cashbook workbook, one sheet per month (Apr..Mar),
    using fn_cashbook_paired_v3. Requires the entry_side migration to be
    deployed - this does not fall back to the pre-entry_side function.
    """
    from database.db_manager import db as _db
    if db is None:
        db = _db

    soc = db._execute(
        "SELECT name, pan_number FROM societies WHERE id=%s",
        (society_id,), fetch_one=True,
    ) or {}
    society_name = soc.get("name", "Society")
    pan = soc.get("pan_number", "")
    asst_year = f"{fy}-{fy+1}"
    filename = f"{filename_prefix}_{fy}-{fy+1}.xlsx"

    fy_start = date(fy, 4, 1)

    # Fixed (2026-08): sourced via fn_cih_balance_asof — the same shared
    # helper fn_cashbook_month_page and fn_account_ledger_fy's CiH branch
    # use, rather than re-deriving the brought_forward lookup inline here
    # a third time. (This also carries forward the earlier
    # is_cash_or_bank -> tab_name='CiH' fix: is_cash_or_bank is never
    # populated by seed.py/migrate.py, so a query against it always
    # resolved to 0 regardless of what brought_forward actually held.)
    #
    # Fixed: was fn_cih_balance_asof(society_id, fy_start - 1 day) —
    # subtracting a day from the FY's own first day crosses into the
    # PRIOR fiscal year (fy_start=2026-04-01 minus a day is 2026-03-31 =
    # FY2025), which this system never seeds a brought_forward row for —
    # each FY's BF is entered directly, with no assumption that the prior
    # FY "closed into" this one. That silently returned 0 for the opening
    # balance instead of the FY's real seeded BF. Same fix as the two SQL
    # functions with this identical pattern (fn_cashbook_paired_v3,
    # fn_cashbook_month_page): call with fy_start itself (always resolves
    # to the FY that owns it) and subtract that day's own transactions
    # back out inline, rather than asking for a date that can cross into
    # a different FY.
    bf_row = db._execute(
        """SELECT fn_cih_balance_asof(%s, %s) - COALESCE((
               SELECT SUM(CASE WHEN t.entry_side = 'Cr' THEN t.amount
                                WHEN t.entry_side = 'Dr' THEN -t.amount
                                ELSE 0 END)
               FROM transactions t
               WHERE t.society_id = %s AND t.status = 'paid' AND t.mode = 'cash'
                 AND t.trx_date = %s
           ), 0) AS bf""",
        (society_id, fy_start, society_id, fy_start), fetch_one=True,
    ) or {}
    opening_balance = float(bf_row.get("bf", 0))

    all_rows = db._execute(
        "SELECT * FROM fn_cashbook_paired_v3(%s, %s, %s, NULL, %s, %s) ORDER BY row_date",
        (society_id, entity_id, entity_role, fy_start, date(fy + 1, 3, 31)),
        fetch_all=True,
    ) or []

    wb = Workbook()
    wb.remove(wb.active)

    month, year = 4, fy
    for _ in range(12):
        month_start = date(year, month, 1)
        month_end = date(year + 1, 1, 1) if month == 12 else date(year, month + 1, 1)
        month_rows = [r for r in all_rows if month_start <= r["row_date"] < month_end]

        ws = wb.create_sheet(title=month_start.strftime("%b"))
        opening_balance = _write_month_sheet(
            ws, month_rows, opening_balance, society_name, pan, asst_year, month_start, filename
        )

        if month == 12:
            month, year = 1, year + 1
        else:
            month += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
