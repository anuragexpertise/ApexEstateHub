#!/usr/bin/env python3
# database/seed.py
"""
First-run seeder for ApexEstateHub.
Handles:
 1. Creating Master Admin if no users exist.
 2. Creating Dummy Society + Users if no societies exist.
"""
import os
import sys
import traceback
import argparse
from pathlib import Path
from openpyxl import load_workbook

# Add project root BEFORE importing project modules
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, PROJECT_ROOT)

from database.db_manager import db



from dotenv import load_dotenv
load_dotenv(override=False)

from werkzeug.security import generate_password_hash

# ── Debug Helper ──────────────────────────────────────────────────────────────

def log_debug(label, query, params=None, result=None, error=None):
    """Prints debug info to help trace issues."""
    print(f"\n🔍 DEBUG: {label}")
    # Truncate long queries for readability
    q_preview = query[:120] + "..." if len(query) > 120 else query
    print(f"   Query: {q_preview}")
    print(f"   Params Type: {type(params).__name__ if params else 'None'}")
    if params:
        # Truncate large dicts for readability
        p_str = str(params)
        if len(p_str) > 150:
            p_str = p_str[:150] + "..."
        print(f"   Params: {p_str}")
    
    if result:
        print(f"   Result: {result}")
    if error:
        print(f"   ERROR: {error}")
        # Uncomment below for full traceback if needed
        # print(f"   Traceback: {''.join(traceback.format_exception(type(error), error, error.__traceback__))}")


# ── Configuration ─────────────────────────────────────────────────────────────

MASTER_EMAIL    = 'master@estatehub.com'
MASTER_PASSWORD = 'Master@2024'

DUMMY_SOCIETY = {
    'name':            'Sunrise Residency',
    'email':           'admin@sunriseresidency.com',
    'phone':           '9876543210',
    'address':         '12, MG Road, Sector 5, Meerut, UP - 250001',
    'secretary_name':  'Ramesh Kumar',
    'secretary_phone': '9876543211',
    'plan':            'Free',
    'plan_validity':   '2025-12-31',
    'arrear_start_date': '2024-04-01',
}

DUMMY_USERS = [
    {
        'role':        'admin',
        'email':       'admin@sunriseresidency.com',
        'password':    'Admin@2024',
        'flat':        None,
        'name':        'Society Admin',
        'description': 'Society Admin',
    },
    {
        'role':        'apartment',
        'email':       'owner@sunriseresidency.com',
        'password':    'Owner@2024',
        'flat':        'A-101',
        'area':        1200,
        'owner_name':  'Rajesh Sharma',
        'mobile':      '9811111111',
        'name':        'Rajesh Sharma',
        'description': 'Apartment Owner (Flat A-101)',
    },
    {
        'role':        'vendor',
        'email':       'vendor@sunriseresidency.com',
        'password':    'Vendor@2024',
        'flat':        None,
        'name':        'Plumbing Services',
        'description': 'Vendor (Plumbing)',
    },
    {
        'role':        'security',
        'email':       'security@sunriseresidency.com',
        'password':    'Security@2024',
        'flat':        None,
        'name':        'Guard Ramu',
        'description': 'Security Staff',
    },
]

# ── 1. Master Admin ───────────────────────────────────────────────────────────

def ensure_master_admin() -> bool:
    """Create master admin if no users exist at all. Returns True if created."""
    
    log_debug("START: ensure_master_admin", "SELECT COUNT(*) FROM users", None)

    # Try empty list [] first, then empty dict {}
    params_list = [[], {}]
    result = None
    last_error = None

    for params in params_list:
        try:
            result = db._execute("SELECT COUNT(*) AS c FROM users", params, fetch_one=True)
            log_debug("SUCCESS: Count Query", "SELECT COUNT(*)", params, result)
            break
        except Exception as e:
            last_error = e
            log_debug("FAIL: Count Query", "SELECT COUNT(*)", params, error=e)

    if last_error and not result:
        print(f"❌  DB error checking users: {last_error}")
        return False

    if result and result.get('c', 0) > 0:
        print(f"✓ Users exist ({result['c']}) — skipping master admin creation.")
        return False

    pw_hash = generate_password_hash(MASTER_PASSWORD)
    insert_params = {
        'email': MASTER_EMAIL,
        'password_hash': pw_hash,
        'role': 'admin',
        'login_method': 'password',
        'is_master_admin': True,
    }
    
    log_debug("START: Insert Master", "INSERT INTO users ...", insert_params)
    try:
        db._execute(
            """
            INSERT INTO users (email, password_hash, role, login_method, is_master_admin)
            VALUES (:email, :password_hash, :role, :login_method, :is_master_admin)
            ON CONFLICT (email) DO UPDATE SET is_master_admin = TRUE
            """,
            insert_params
        )
        log_debug("SUCCESS: Insert Master", "INSERT INTO users ...", insert_params, result="Inserted")
        
        print()
        print("  ┌─────────────────────────────────────────────┐")
        print("  │          Master Admin Created  ✓             │")
        print("  ├─────────────────────────────────────────────┤")
        print(f"  │  Email    : {MASTER_EMAIL:<33}│")
        print(f"  │  Password : {MASTER_PASSWORD:<33}│")
        print("  │  ⚠  Change this password after first login! │")
        print("  └─────────────────────────────────────────────┘")
        print()
        return True
    except Exception as e:
        log_debug("FAIL: Insert Master", "INSERT INTO users ...", insert_params, error=e)
        print(f"❌  Failed to create master admin: {e}")
        return False

# ── 2. Dummy Society + Users ──────────────────────────────────────────────────

def ensure_dummy_data():
    """If no societies exist, prompt user and optionally seed dummy data."""
    
    log_debug("START: ensure_dummy_data", "SELECT COUNT(*) FROM societies", None)

    params_list = [[], {}]
    result = None
    last_error = None

    for params in params_list:
        try:
            result = db._execute("SELECT COUNT(*) AS c FROM societies", params, fetch_one=True)
            log_debug("SUCCESS: Count Societies", "SELECT COUNT(*)", params, result)
            break
        except Exception as e:
            last_error = e
            log_debug("FAIL: Count Societies", "SELECT COUNT(*)", params, error=e)

    if last_error and not result:
        print(f"❌  DB error checking societies: {last_error}")
        return

    if result and result.get('c', 0) > 0:
        print(f"✓ Societies exist ({result['c']}) — skipping dummy data.")
        return

    print()
    print("  ℹ  No societies found (first run).")
    print("  Would you like to create dummy data?")
    print("  This will create:")
    print("    • 1 society  : Sunrise Residency")
    print("    • 1 admin    : admin@sunriseresidency.com  / Admin@2024")
    print("    • 1 owner    : owner@sunriseresidency.com  / Owner@2024  (Flat A-101)")
    print("    • 1 vendor   : vendor@sunriseresidency.com / Vendor@2024")
    print("    • 1 security : security@sunriseresidency.com / Security@2024")
    print()

    try:
        answer = input("  Create dummy data? [y/N]: ").strip().lower()
    except (EOFError, KeyboardInterrupt):
        answer = 'n'

    if answer != 'y':
        print("  Skipped. You can create a society by logging in as master admin.")
        return

    _seed_dummy(db)

def _seed_dummy(db):
    society_id = None

    # 1. Check existence
    check_params = {'name': DUMMY_SOCIETY['name']}
    log_debug("START: Check Society", "SELECT id FROM societies WHERE name = :name", check_params)
    try:
        existing = db._execute(
            "SELECT id FROM societies WHERE name = :name",
            check_params,
            fetch_one=True
        )
        log_debug("RESULT: Check Society", "SELECT id ...", check_params, existing)
        
        if existing and existing.get('id'):
            society_id = existing['id']
            print(f"  ✓ Society already exists (id={society_id}): {DUMMY_SOCIETY['name']}")
        else:
            # 2. Insert
            insert_params = {
                'name': DUMMY_SOCIETY['name'],
                'email': DUMMY_SOCIETY['email'],
                'phone': DUMMY_SOCIETY['phone'],
                'address': DUMMY_SOCIETY['address'],
                'secretary_name': DUMMY_SOCIETY['secretary_name'],
                'secretary_phone': DUMMY_SOCIETY['secretary_phone'],
                'plan': DUMMY_SOCIETY['plan'],
                'plan_validity': DUMMY_SOCIETY['plan_validity'],
                'arrear_start_date': DUMMY_SOCIETY['arrear_start_date'],
            }
            log_debug("START: Insert Society", "INSERT INTO societies ...", insert_params)
            try:
                result = db._execute(
                    """
                    INSERT INTO societies
                        (name, email, phone, address, secretary_name, secretary_phone,
                         plan, plan_validity, arrear_start_date)
                    VALUES (:name, :email, :phone, :address, :secretary_name, :secretary_phone,
                            :plan, :plan_validity, :arrear_start_date)
                    RETURNING id
                    """,
                    insert_params,
                    fetch_one=True
                )
                log_debug("RESULT: Insert Society", "INSERT INTO societies ...", insert_params, result)
                
                if result and result.get('id'):
                    society_id = result['id']
                    print(f"  ✓ Society created (id={society_id}): {DUMMY_SOCIETY['name']}")
                else:
                    # Fallback: fetch again (in case RETURNING failed but insert succeeded)
                    log_debug("FALLBACK: Fetch Society", "SELECT id FROM societies WHERE name = :name", check_params)
                    retry = db._execute(
                        "SELECT id FROM societies WHERE name = :name",
                        check_params,
                        fetch_one=True
                    )
                    log_debug("RESULT: Fallback Fetch", "SELECT id ...", check_params, retry)
                    if retry and retry.get('id'):
                        society_id = retry['id']
                        print(f"  ✓ Society found after insert (id={society_id}).")
                    else:
                        print(f"  ❌  Failed to create or retrieve society.")
                        return

            except Exception as e:
                log_debug("FAIL: Insert Society", "INSERT INTO societies ...", insert_params, error=e)
                # Handle race condition: duplicate key error
                if 'duplicate key' in str(e).lower() or 'UniqueViolation' in str(type(e)):
                    log_debug("RETRY: Duplicate Key", "SELECT id FROM societies WHERE name = :name", check_params)
                    retry = db._execute(
                        "SELECT id FROM societies WHERE name = :name",
                        check_params,
                        fetch_one=True
                    )
                    log_debug("RESULT: Retry Fetch", "SELECT id ...", check_params, retry)
                    if retry and retry.get('id'):
                        society_id = retry['id']
                        print(f"  ✓ Society already existed (race condition handled, id={society_id}).")
                    else:
                        print(f"  ❌  Society exists but could not retrieve ID: {e}")
                        return
                else:
                    print(f"  ❌  Society creation failed: {e}")
                    return

    except Exception as e:
        log_debug("FAIL: Check Society", "SELECT id ...", check_params, error=e)
        print(f"  ❌  Society check failed: {e}")
        return

    if not society_id:
        print("  ❌  Could not determine society ID. Aborting user creation.")
        return

    # 3. Create users
    for u in DUMMY_USERS:
        pw_hash = generate_password_hash(u['password'])
        user_params = {
            'society_id': society_id,
            'email': u['email'],
            'password_hash': pw_hash,
            'role': u['role'],
            'login_method': 'password'
        }
        log_debug(f"START: Insert User ({u['role']})", "INSERT INTO users ...", user_params)
        try:
            user = db._execute(
                """
                INSERT INTO users (society_id, email, password_hash, role, login_method)
                VALUES (:society_id, :email, :password_hash, :role, :login_method)
                ON CONFLICT (email) DO NOTHING
                RETURNING id
                """,
                user_params,
                fetch_one=True
            )
            log_debug(f"RESULT: Insert User ({u['role']})", "INSERT INTO users ...", user_params, user)
            
            if not user:
                print(f"  ⚠  {u['description']} already exists — skipped.")
                continue

            user_id = user['id']

            if u['role'] == 'apartment' and u.get('flat'):
                apt_params = {
                    'society_id': society_id,
                    'flat_number': u['flat'],
                    'owner_name': u.get('owner_name', u['name']),
                    'mobile': u.get('mobile', ''),
                    'apartment_size': u.get('area', 1000),
                    'active': True
                }
                log_debug(f"START: Insert Apartment", "INSERT INTO apartments ...", apt_params)
                apt = db._execute(
                    """
                    INSERT INTO apartments
                        (society_id, flat_number, owner_name, mobile, apartment_size, active)
                    VALUES (:society_id, :flat_number, :owner_name, :mobile, :apartment_size, :active)
                    ON CONFLICT (society_id, flat_number) DO NOTHING
                    RETURNING id
                    """,
                    apt_params,
                    fetch_one=True
                )
                log_debug(f"RESULT: Insert Apartment", "INSERT INTO apartments ...", apt_params, apt)
                if apt:
                    update_params = {'linked_id': apt['id'], 'user_id': user_id}
                    log_debug("UPDATE: Link User", "UPDATE users SET linked_id ...", update_params)
                    db._execute(
                        "UPDATE users SET linked_id=:linked_id WHERE id=:user_id",
                        update_params
                    )
                    log_debug("SUCCESS: Link User", "UPDATE users ...", update_params, result="Done")

            print(f"  ✓ {u['description']:<30} → {u['email']}  /  {u['password']}")

        except Exception as e:
            log_debug(f"FAIL: Insert User ({u['role']})", "INSERT INTO users ...", user_params, error=e)
            print(f"  ❌  {u['description']} failed: {e}")

    print()
    print("  ┌─────────────────────────────────────────────────────────┐")
    print("  │              Dummy Data Created  ✓                       │")
    print("  │  Login at http://127.0.0.1:8050                          │")
    print("  │  Master : master@estatehub.com     / Master@2024         │")
    print("  │  Admin  : admin@sunriseresidency.com / Admin@2024        │")
    print("  │  Owner  : owner@sunriseresidency.com / Owner@2024        │")
    print("  │  Vendor : vendor@sunriseresidency.com / Vendor@2024      │")
    print("  │  Sec    : security@sunriseresidency.com / Security@2024  │")
    print("  └─────────────────────────────────────────────────────────┘")
# ── 3. Estate Account Seeder (Excel Import) ──────────────────────────────────

def ensure_accounts_table():
    """
    Creates accounts table if not present.
    Safe to run repeatedly.
    """

    create_sql = """
        CREATE TABLE IF NOT EXISTS accounts (
        id SERIAL PRIMARY KEY,
        society_id INT NOT NULL REFERENCES societies (id) ON DELETE CASCADE,
        name VARCHAR(20) NOT NULL,
        tab_name VARCHAR(10),
        header VARCHAR(50),
        parent_account_id INT NOT NULL REFERENCES accounts (id) DEFAULT 1,
        drcr_account VARCHAR(2) CHECK (drcr_account IN ('Dr', 'Cr')) NOT NULL,
        has_bf BOOLEAN DEFAULT FALSE,
        drcr_bf VARCHAR(2) CHECK (drcr_bf IN ('Dr', 'Cr')) NOT NULL,
        bf_amount DECIMAL(12, 2) DEFAULT 0.00,
        depreciation_percent DECIMAL(5, 2) DEFAULT 0.00,
        created_at TIMESTAMP DEFAULT NOW(),
        CONSTRAINT uq_account_society_name UNIQUE (society_id, name),
        CONSTRAINT uq_account_society_id UNIQUE (society_id, id)
        );
        """

    try:
        db._execute(create_sql, {})
        print("  ✓ accounts table ready")
    except Exception as e:
        print(f"  ❌ Failed creating accounts table: {e}")
        raise


def import_estate_accounts(force=False):
    """
    Imports EstateAcc.xlsx into accounts table.

    Runs only when:
        migrate.py executed with --force
    """

    print()
    print("── Estate Account Import ───────────────────────────────────")

    ensure_accounts_table()

    project_root = Path(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    excel_path = project_root / "EstateAcc.xlsx"

    if not excel_path.exists():
        print(f"  ⚠ Excel file not found: {excel_path}")
        return

    try:
        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active
    except Exception as e:
        print(f"  ❌ Failed opening workbook: {e}")
        return

    imported = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):

        if not row:
            continue

        id = row[0]

        # skip blank rows
        if id in (None, ''):
            continue

        try:
            params = {
                'id': int(id) if id is not None else None,
                'name': row[1],
                'tab_name': row[2],
                'header': row[3],
                'parent_account_id': row[4],
                'drcr_bf': bool(row[5]) if row[5] is not None else None,
                'has_bf': bool(row[6]) if row[6] is not None else None,
                'bf_amount': float(row[7] or 0),
                'depreciation_percent': float(row[8] or 0),
                'drcr_ac': bool(row[9]) if row[9] is not None else None,
            }

            existing = db._execute(
                """
                SELECT id
                FROM accounts
                WHERE id = :id
                """,
                {'id': params['id']},
                fetch_one=True
            )

            if existing and not force:
                skipped += 1
                continue

            if existing and force:
                db._execute(
                    """
                    UPDATE accounts
                    SET
                        name = :name,
                        tab_name     = :tab_name,
                        header  = :header,
                        parent_account_id    = :parent_account_id,
                        drcr_bf     = :drcr_bf,
                        has_bf        = :has_bf,
                        bf_amount    = :bf_amount,
                        depreciation_percent = :depreciation_percent,
                        drcr_ac     = :drcr_ac
                    WHERE id = :id
                    """,
                    params
                )
            else:
                db._execute(
                    """
                    INSERT INTO accounts (
                        id,
                        name,
                        tab_name,
                        header,
                        parent_account_id,
                        drcr_bf,
                        has_bf,
                        bf_amount,
                        depreciation_percent,
                        drcr_ac
                    )
                    VALUES (
                        :id,
                        :name,
                        :tab_name,
                        :header,
                        :parent_account_id,
                        :drcr_bf,
                        :has_bf,
                        :bf_amount,
                        :depreciation_percent,
                        :drcr_ac
                    )
                    """,
                    params
                )

            imported += 1

        except Exception as e:
            skipped += 1
            print(f"  ⚠ Skipped account row {id}: {e}")

    print(f"  ✓ Imported : {imported}")
    print(f"  ✓ Skipped  : {skipped}")

def run(force=False):
    print()
    print("── Seed check (DEBUG MODE) ───────────────────────────────────────────────")

    ensure_master_admin()
    ensure_dummy_data()

    # Import Excel account master only on forced migration
    if force:
        import_estate_accounts(force=True)

    print("─────────────────────────────────────────────────────────────")

if __name__ == '__main__':

    parser = argparse.ArgumentParser()
    parser.add_argument('--force', action='store_true')
    args = parser.parse_args()

    run(force=args.force)