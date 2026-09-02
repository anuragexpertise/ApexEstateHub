# database/qr_reissue_export.py
"""
QR Reissue Audit Log Excel Generator — EstateHub
=================================================
Exports qr_reissue_log — the append-only record of every admin-initiated
QR revoke/reissue (apartment/vendor/security/admin/patrol_location),
written exclusively by app/services/qr_service.py's revoke_and_reissue.

One sheet, one row per reissue event: when, which entity, why, who did it,
old nonce -> new nonce.
"""

from __future__ import annotations
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


_FONT_BODY = Font(name="Arial", size=9)
_FONT_HEADER = Font(name="Arial", size=9, bold=True)
_FONT_TITLE = Font(name="Arial", size=10, bold=True)

_FILL_HEADER = PatternFill("solid", fgColor="D9E1F2")

_ALIGN_C = Alignment(horizontal="center", vertical="center")
_ALIGN_L = Alignment(horizontal="left", vertical="center")

_THIN = Side(style="thin")
_BORDER_ALL = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_FMT_DT = "DD-MMM-YYYY HH:MM"

_COL_WIDTHS = {
    "A": 18, "B": 10, "C": 10, "D": 28, "E": 10,
    "F": 10, "G": 12, "H": 22,
}

_ROLE_LABELS = {
    "APT": "Apartment", "VND": "Vendor", "SEC": "Security",
    "ADM": "Admin", "PTL": "Patrol Location",
}
_REASON_LABELS = {
    "lost": "Lost", "theft": "Theft", "mutilated": "Mutilated",
    "request": "Requested", "other": "Other",
}


def _write_log_sheet(ws, rows: list[dict], society_label: str) -> None:
    ws.cell(row=1, column=1, value=f"QR Reissue Audit Log — {society_label}")
    ws.cell(row=1, column=1).font = _FONT_TITLE

    headers = [
        "Date/Time", "Role", "Entity ID", "Entity",
        "Reason", "Old Code", "New Code", "Revoked By",
    ]
    for col, hdr in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=hdr)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_C
        cell.border = _BORDER_ALL

    r = 3
    for row in rows:
        created_at = row.get("created_at")
        role_code = row.get("role_code")
        values = [
            created_at,
            _ROLE_LABELS.get(role_code, role_code),
            row.get("entity_id"),
            row.get("entity_label") or "—",
            _REASON_LABELS.get(row.get("reason"), row.get("reason")),
            row.get("old_nonce") or "—",
            row.get("new_nonce"),
            row.get("actor_name") or f"user #{row.get('actor_user_id')}",
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = _FONT_BODY
            cell.border = _BORDER_ALL
            if col == 1 and isinstance(val, datetime):
                cell.number_format = _FMT_DT
                cell.alignment = _ALIGN_L
            elif col == 4:
                cell.alignment = _ALIGN_L
            else:
                cell.alignment = _ALIGN_C
        r += 1

    if not rows:
        ws.cell(row=r, column=1, value="No reissue events recorded.").font = _FONT_BODY


def generate_qr_reissue_log_excel(
    db,
    society_id: int,
    society_label: str = None,
    date_from=None,
    date_to=None,
) -> bytes:
    """
    Builds the QR reissue audit log workbook for one society, optionally
    bounded by [date_from, date_to] (inclusive; either or both may be
    None for an unbounded log). actor_name is resolved via a join against
    users so the export reads as a name rather than a bare user id.
    """
    from database.db_manager import db as _db
    if db is None:
        db = _db

    conditions = ["l.society_id = %s"]
    params = [society_id]
    if date_from:
        conditions.append("l.created_at >= %s")
        params.append(date_from)
    if date_to:
        conditions.append("l.created_at <= %s")
        params.append(date_to)

    rows = db._execute(
        f"""SELECT l.*, u.name AS actor_name
              FROM qr_reissue_log l
              LEFT JOIN users u ON u.id = l.actor_user_id
             WHERE {' AND '.join(conditions)}
             ORDER BY l.created_at DESC""",
        tuple(params), fetch_all=True,
    ) or []

    if not society_label:
        soc = db._execute(
            "SELECT name FROM societies WHERE id=%s", (society_id,), fetch_one=True
        ) or {}
        society_label = soc.get("name") or f"Society #{society_id}"

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title="QR Reissue Log")
    for col, width in _COL_WIDTHS.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[2].height = 22

    _write_log_sheet(ws, rows, society_label)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
