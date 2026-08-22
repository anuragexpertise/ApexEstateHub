# database/tds_export.py
"""
TDS Return (Form 26Q) Excel Generator — EstateHub
===================================================
Produces a structured quarterly TDS-deduction report an society's CA can
transcribe into the government portal (26Q). This is NOT a TRACES e-filing
integration — it is the structured-data boundary the rest of this project
draws: correct rows, flagged for filing-blocking problems (missing PAN),
no government-API/file-format coupling.

Data source: fn_tds_summary_fy(p_society_id, p_fy, p_quarter) — one row per
TDS-deducted payment (26Q wants per-transaction deduction records with
dates, not vendor totals).

Two sheets:
  * "TDS Deductions" — per-transaction rows (vendor_name, vendor_pan,
        tds_section, gross_amount_paid, tds_deducted, net_paid,
        payment_date). Rows for vendors with no PAN are highlighted red —
        a missing PAN is filing-blocking and must surface loudly.
  * "Vendor Summary" — one row per (vendor, section, FY): cumulative
        gross / tds / net + a no-PAN flag — a human cross-check against
        Phase 4.2's annual-aggregate threshold tracking.

[-WFLAG — PROFESSIONAL REVIEW- The 26Q column layout/order is confirmed
against whichever return-filing software the society's CA actually uses
before treating this sheet as final. "Structured and correct" and
"matches what gets copy-pasted into the filing tool" are two bars.]

Explicitly out of scope: TRACES/26Q e-filing, TAN registration, government
file-format validation.
"""

from __future__ import annotations
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


_FONT_BODY = Font(name="Arial", size=9)
_FONT_HEADER = Font(name="Arial", size=9, bold=True)
_FONT_TITLE = Font(name="Arial", size=10, bold=True)
_FONT_TOTAL = Font(name="Arial", size=9, bold=True)

_FILL_HEADER = PatternFill("solid", fgColor="D9E1F2")
_FILL_NOPAN = PatternFill("solid", fgColor="F8CBAD")
_FILL_TOTAL = PatternFill("solid", fgColor="E2EFDA")

_ALIGN_C = Alignment(horizontal="center", vertical="center")
_ALIGN_L = Alignment(horizontal="left", vertical="center")
_ALIGN_R = Alignment(horizontal="right", vertical="center")

_THIN = Side(style="thin")
_BORDER_ALL = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_FMT_AMT = '#,##0.00;[Red](#,##0.00);"-"'
_FMT_DATE = "DD-MMM-YY"

_COL_WIDTHS = {
    "A": 22, "B": 14, "C": 10, "D": 16, "E": 14, "F": 14, "G": 12, "H": 8,
}

_SUM_COL_WIDTHS = {
    "A": 22, "B": 14, "C": 10, "D": 16, "E": 14, "F": 14, "G": 12, "H": 8, "I": 10,
}


def _apply_header(ws, row: int, widths: dict) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[row].height = 22


def _quarter_label(fy: int, quarter: int) -> str:
    months = {1: "Apr-Jun", 2: "Jul-Sep", 3: "Oct-Dec", 4: "Jan-Mar"}
    return f"FY {fy}-{fy + 1} Q{quarter} ({months.get(quarter, '')})"


def _write_deduction_sheet(ws, rows: list[dict], fy: int, quarter: int) -> int:
    ws.cell(row=1, column=1, value=f"TDS Deductions — {_quarter_label(fy, quarter)}")
    ws.cell(row=1, column=1).font = _FONT_TITLE

    headers = ["Vendor Name", "PAN", "TDS Section", "Gross Paid",
               "TDS Deducted", "Net Paid", "Payment Date", "No PAN"]
    for col, hdr in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=hdr)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_C
        cell.border = _BORDER_ALL

    r = 3
    for row in rows:
        no_pan = bool(row.get("no_pan"))
        values = [
            row.get("vendor_name"), row.get("vendor_pan"), row.get("tds_section"),
            float(row.get("gross_amount_paid", 0) or 0),
            float(row.get("tds_deducted", 0) or 0),
            float(row.get("net_paid", 0) or 0),
            row.get("payment_date"), "YES" if no_pan else "",
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = _FONT_BODY
            cell.border = _BORDER_ALL
            if no_pan:
                cell.fill = _FILL_NOPAN
            if col in (4, 5, 6):
                cell.alignment = _ALIGN_R
                cell.number_format = _FMT_AMT
            elif col == 7 and val is not None:
                cell.alignment = _ALIGN_C
                cell.number_format = _FMT_DATE
            else:
                cell.alignment = _ALIGN_L
        r += 1

    # Totals row.
    t_gross = sum(float(x.get("gross_amount_paid", 0) or 0) for x in rows)
    t_tds = sum(float(x.get("tds_deducted", 0) or 0) for x in rows)
    t_net = sum(float(x.get("net_paid", 0) or 0) for x in rows)
    totals = ["TOTAL", "", "", t_gross, t_tds, t_net, ""]
    for col, val in enumerate(totals, start=1):
        cell = ws.cell(row=r, column=col, value=val)
        cell.font = _FONT_TOTAL
        cell.fill = _FILL_TOTAL
        cell.border = _BORDER_ALL
        if col in (4, 5, 6):
            cell.alignment = _ALIGN_R
            cell.number_format = _FMT_AMT
        else:
            cell.alignment = _ALIGN_L
    return r


def _write_summary_sheet(ws, rows: list[dict], fy: int, quarter: int) -> None:
    ws.cell(row=1, column=1, value=f"TDS Vendor Summary — {_quarter_label(fy, quarter)}")
    ws.cell(row=1, column=1).font = _FONT_TITLE

    headers = ["Vendor Name", "PAN", "TDS Section", "Gross Paid",
               "TDS Deducted", "Net Paid", "No PAN", "No-PAN Count", "Bills"]
    for col, hdr in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=hdr)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_C
        cell.border = _BORDER_ALL

    # Vendor-level aggregation: same query, grouped by vendor+section.
    agg: dict[tuple, dict] = {}
    for row in rows:
        key = (
            row.get("vendor_name"),
            row.get("vendor_pan"),
            row.get("tds_section"),
        )
        bucket = agg.setdefault(key, {
            "gross": 0.0, "tds": 0.0, "net": 0.0,
            "no_pan": False, "bills": 0,
        })
        bucket["gross"] += float(row.get("gross_amount_paid", 0) or 0)
        bucket["tds"] += float(row.get("tds_deducted", 0) or 0)
        bucket["net"] += float(row.get("net_paid", 0) or 0)
        bucket["no_pan"] = bucket["no_pan"] or bool(row.get("no_pan"))
        bucket["bills"] += 1

    r = 3
    for (name, pan, section), b in agg.items():
        values = [name, pan, section, b["gross"], b["tds"], b["net"],
                  "YES" if b["no_pan"] else "", b["bills"] if b["no_pan"] else "", b["bills"]]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = _FONT_BODY
            cell.border = _BORDER_ALL
            if b["no_pan"]:
                cell.fill = _FILL_NOPAN
            if col in (4, 5, 6):
                cell.alignment = _ALIGN_R
                cell.number_format = _FMT_AMT
            elif col in (8, 9):
                cell.alignment = _ALIGN_C
            else:
                cell.alignment = _ALIGN_L
        r += 1


def generate_tds_summary_excel(
    db,
    society_id: int,
    fy: int,
    quarter: int = 1,
    filename_prefix: str = "TDS26Q",
) -> bytes:
    """Builds the quarterly 26Q TDS workbook for one society.

    fy       — FY START year (e.g. 2026 for FY 1-Apr-2026..31-Mar-2027)
    quarter  — 1..4 (Q1 Apr-Jun ... Q4 Jan-Mar, straddling the FY boundary)
    """
    from database.db_manager import db as _db
    if db is None:
        db = _db

    rows = db._execute(
        "SELECT * FROM fn_tds_summary_fy(%s,%s,%s)",
        (society_id, str(fy), quarter), fetch_all=True,
    ) or []

    no_pan_count = sum(1 for x in rows if x.get("no_pan"))

    wb = Workbook()
    wb.remove(wb.active)

    ws_ded = wb.create_sheet(title="TDS Deductions")
    _apply_header(ws_ded, 2, _COL_WIDTHS)
    last_row = _write_deduction_sheet(ws_ded, rows, fy, quarter)
    if no_pan_count:
        note_cell = ws_ded.cell(
            row=last_row + 2, column=1,
            value=(f"{no_pan_count} row(s) have NO PAN on file — "
                   "filing-blocked until captured."),
        )
        note_cell.font = Font(name="Arial", size=9, bold=True, color="C00000")

    ws_sum = wb.create_sheet(title="Vendor Summary")
    _apply_header(ws_sum, 2, _SUM_COL_WIDTHS)
    _write_summary_sheet(ws_sum, rows, fy, quarter)

    filename = f"{filename_prefix}_FY{fy}-Q{quarter}.xlsx"
    # Stash filename on the workbook so the dispatcher can pick it up.
    wb._tds_filename = filename  # type: ignore[attr-defined]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
