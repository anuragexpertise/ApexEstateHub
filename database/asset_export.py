# database/asset_export.py
"""
Fixed Asset Register Excel Generator — EstateHub
"""

from __future__ import annotations
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

_FONT_BODY = Font(name="Arial", size=9)
_FONT_HEADER = Font(name="Arial", size=9, bold=True)
_FONT_TITLE = Font(name="Arial", size=10, bold=True)
_FONT_TOTAL = Font(name="Arial", size=9, bold=True)

_FILL_HEADER = PatternFill("solid", fgColor="D9E1F2")
_FILL_TOTAL = PatternFill("solid", fgColor="E2EFDA")

_ALIGN_C = Alignment(horizontal="center", vertical="center")
_ALIGN_L = Alignment(horizontal="left", vertical="center")
_ALIGN_R = Alignment(horizontal="right", vertical="center")

_THIN = Side(style="thin")
_BORDER_ALL = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_FMT_AMT = '#,##0.00;[Red](#,##0.00);"-"'
_FMT_DATE = "DD-MMM-YYYY"

def _apply_header(ws, row: int, widths: dict) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[row].height = 22

def _write_block_summary_sheet(ws, rows: list[dict], fy: int) -> None:
    ws.cell(row=1, column=1, value=f"Fixed Asset Register — Block Summary FY {fy}-{fy+1}")
    ws.cell(row=1, column=1).font = _FONT_TITLE

    headers = [
        "Block / Account Name", "Dep Rate %", "Opening WDV",
        "Additions (>=180 days use)", "Additions (<180 days use)", "Deductions / Sales",
        "Depreciation Charge", "Closing WDV", "STCG u/s 50", "STCL u/s 50",
    ]
    widths = {"A": 25, "B": 12, "C": 15, "D": 20, "E": 20, "F": 18, "G": 18, "H": 18, "I": 14, "J": 14}
    _apply_header(ws, 2, widths)

    for col, hdr in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=hdr)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_C
        cell.border = _BORDER_ALL

    r = 3
    t_op, t_a1, t_a2, t_ded, t_dep, t_cl, t_stcg, t_stcl = 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0
    for row in rows:
        acc_name = row.get("account_name", "")
        pct = float(row.get("depreciation_percent") or 0)
        op_wdv = float(row.get("opening_wdv") or 0)
        a1 = float(row.get("additions_first_half") or 0)
        a2 = float(row.get("additions_second_half") or 0)
        ded = float(row.get("deductions") or 0)
        dep = float(row.get("depreciation_charge") or 0)
        cl_wdv = float(row.get("closing_wdv") or 0)
        stcg = float(row.get("stcg_u_s_50") or 0)
        stcl = float(row.get("stcl_u_s_50") or 0)

        vals = [acc_name, pct, op_wdv, a1, a2, ded, dep, cl_wdv, stcg, stcl]
        for col, val in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = _FONT_BODY
            cell.border = _BORDER_ALL
            if col == 1:
                cell.alignment = _ALIGN_L
            elif col == 2:
                cell.alignment = _ALIGN_C
                cell.number_format = '0.00'
            else:
                cell.alignment = _ALIGN_R
                cell.number_format = _FMT_AMT
        r += 1
        t_op += op_wdv
        t_a1 += a1
        t_a2 += a2
        t_ded += ded
        t_dep += dep
        t_cl += cl_wdv
        t_stcg += stcg
        t_stcl += stcl

    if rows:
        totals = ["TOTAL", "", t_op, t_a1, t_a2, t_ded, t_dep, t_cl, t_stcg, t_stcl]
        for col, val in enumerate(totals, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = _FONT_TOTAL
            cell.fill = _FILL_TOTAL
            cell.border = _BORDER_ALL
            if col == 1: cell.alignment = _ALIGN_L
            elif col > 2:
                cell.alignment = _ALIGN_R
                cell.number_format = _FMT_AMT
        if t_stcg or t_stcl:
            r += 2
            note = ws.cell(row=r, column=1,
                value="STCG/STCL u/s 50 are separate short-term capital gains/losses — report them in the capital gains "
                      "schedule of the FY return, not as ordinary business income/expense.")
            note.font = Font(name="Arial", size=8, italic=True, color="C0392B")

def _write_asset_list_sheet(ws, rows: list[dict], fy: int) -> None:
    ws.cell(row=1, column=1, value=f"Fixed Asset Details FY {fy}-{fy+1}")
    ws.cell(row=1, column=1).font = _FONT_TITLE

    headers = [
        "Block / Account Name", "Asset Name", "Serial Number",
        "Purchase Date", "Purchase Value", "Status", "Sale Value",
    ]
    widths = {"A": 25, "B": 25, "C": 20, "D": 15, "E": 18, "F": 12, "G": 18}
    _apply_header(ws, 2, widths)

    for col, hdr in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=hdr)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_C
        cell.border = _BORDER_ALL

    r = 3
    for row in rows:
        acc = row.get("account_name", "")
        name = row.get("asset_name", "")
        sno = row.get("asset_sno", "")
        pdate = row.get("purchase_date")
        pval = float(row.get("purchase_value") or 0)
        disp = bool(row.get("disposed"))
        sval = float(row.get("sale_value") or 0)

        status = "Disposed" if disp else "Active"
        vals = [acc, name, sno, pdate, pval, status, sval if disp else None]
        for col, val in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = _FONT_BODY
            cell.border = _BORDER_ALL
            if col in (1, 2, 3, 6):
                cell.alignment = _ALIGN_L
            elif col == 4:
                cell.alignment = _ALIGN_C
                if val: cell.number_format = _FMT_DATE
            else:
                cell.alignment = _ALIGN_R
                if val is not None: cell.number_format = _FMT_AMT
        r += 1

def generate_fixed_asset_register_excel(db, society_id: int, fy: int) -> bytes:
    if not db:
        from database.db_query import get_connection
        db = get_connection()
    
    blocks = db._execute("SELECT * FROM fn_fixed_asset_register_fy(%s, %s)", (society_id, fy), fetch_all=True)
    assets = db._execute("SELECT * FROM fn_fixed_assets_list_fy(%s, %s)", (society_id, fy), fetch_all=True)

    wb = Workbook()
    ws_blocks = wb.active
    ws_blocks.title = "Block Summary"
    _write_block_summary_sheet(ws_blocks, blocks, fy)

    ws_assets = wb.create_sheet(title="Asset Details")
    _write_asset_list_sheet(ws_assets, assets, fy)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
