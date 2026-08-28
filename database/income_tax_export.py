# database/income_tax_export.py
"""
Income Tax (Principle of Mutuality) Summary Excel Generator — EstateHub
=========================================================================
Produces a structured mutual-vs-non-mutual income/expense breakup for one
financial year, for the society's CA to determine taxable income.

Under the principle of mutuality (see Bangalore Club v. CIT and settled RWA
case law), member-contribution income (maintenance, sinking/repair fund
contributions, etc.) is exempt as mutual income, but investment income
(bank/FD interest) and non-member income (guest event tickets, advertisement
income, etc.) are NOT — they're taxable under normal provisions regardless
of mutuality. accounts.mutuality_nature ('mutual'/'non_mutual') tags which
accounts fall into which bucket; this report sums transactions by that tag.

One sheet:
  * "Mutuality Summary" — Income and Expense, each split mutual/non_mutual,
    with a taxable-income-estimate line (non_mutual income − non_mutual
    expense) for the CA to start from.

Data source: fn_income_tax_summary_fy(p_society_id, p_fy)
"""

from __future__ import annotations
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


_FONT_BODY = Font(name="Arial", size=9)
_FONT_HEADER = Font(name="Arial", size=9, bold=True)
_FONT_TITLE = Font(name="Arial", size=10, bold=True)
_FONT_TOTAL = Font(name="Arial", size=9, bold=True)

_FILL_HEADER = PatternFill("solid", fgColor="D9E1F2")
_FILL_TOTAL = PatternFill("solid", fgColor="E2EFDA")

_ALIGN_C = Alignment(horizontal="center", vertical="center")
_ALIGN_L = Alignment(horizontal="left", vertical="center")
_ALIGN_R = Alignment(horizontal="right", vertical="center")

_THIN = Side(style="thin")
_BORDER_ALL = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_FMT_AMT = '#,##0.00;[Red](#,##0.00);"-"'

_COL_WIDTHS = {"A": 22, "B": 18, "C": 20, "D": 20}


def _apply_header(ws, row: int, widths: dict) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[row].height = 22


def _write_mutuality_sheet(ws, rows: list[dict], fy: int) -> None:
    ws.cell(row=1, column=1, value=f"Income Tax — Mutuality Summary FY {fy}-{fy+1}")
    ws.cell(row=1, column=1).font = _FONT_TITLE

    headers = ["Category", "Nature", "Total Amount"]
    for col, hdr in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=hdr)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_C
        cell.border = _BORDER_ALL

    totals = {
        ("Income", "mutual"): 0.0, ("Income", "non_mutual"): 0.0,
        ("Expense", "mutual"): 0.0, ("Expense", "non_mutual"): 0.0,
    }
    r = 3
    for row in rows:
        category = row.get("category") or ""
        nature = row.get("nature") or ""
        amount = float(row.get("total_amount", 0) or 0)
        totals[(category, nature)] = totals.get((category, nature), 0.0) + amount

        values = [category, nature.replace("_", " ").title(), amount]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = _FONT_BODY
            cell.border = _BORDER_ALL
            if col == 3:
                cell.alignment = _ALIGN_R
                cell.number_format = _FMT_AMT
            else:
                cell.alignment = _ALIGN_L
        r += 1

    r += 1
    taxable_estimate = (
        totals[("Income", "non_mutual")] - totals[("Expense", "non_mutual")]
    )
    summary_rows = [
        ("Total Mutual Income (exempt)", totals[("Income", "mutual")]),
        ("Total Non-Mutual Income (taxable)", totals[("Income", "non_mutual")]),
        ("Total Non-Mutual Expense", totals[("Expense", "non_mutual")]),
        ("Estimated Taxable Income (Non-Mutual Income − Expense)", taxable_estimate),
    ]
    for label, val in summary_rows:
        cell_a = ws.cell(row=r, column=1, value=label)
        cell_a.font = _FONT_TOTAL
        cell_a.fill = _FILL_TOTAL
        cell_a.alignment = _ALIGN_L
        cell_a.border = _BORDER_ALL
        ws.cell(row=r, column=2).fill = _FILL_TOTAL
        ws.cell(row=r, column=2).border = _BORDER_ALL
        cell_c = ws.cell(row=r, column=3, value=val)
        cell_c.font = _FONT_TOTAL
        cell_c.fill = _FILL_TOTAL
        cell_c.alignment = _ALIGN_R
        cell_c.number_format = _FMT_AMT
        cell_c.border = _BORDER_ALL
        r += 1

    note = ws.cell(
        row=r + 1, column=1,
        value=(
            "Note: figures above are estimates for the society's CA to review — "
            "consult a qualified Chartered Accountant before filing. Mutual income "
            "(member contributions) is generally exempt under the Principle of "
            "Mutuality; non-mutual income (e.g. bank/FD interest, non-member "
            "receipts) is taxable under normal provisions regardless of mutuality."
        ),
    )
    note.font = Font(name="Arial", size=8, italic=True)
    note.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=4)
    ws.row_dimensions[r + 1].height = 45


def generate_income_tax_summary_excel(
    db,
    society_id: int,
    fy: int,
    filename_prefix: str = "MutualitySummary",
) -> bytes:
    """Builds the Income Tax mutuality summary workbook for one society FY."""
    from database.db_manager import db as _db
    if db is None:
        db = _db

    rows = db._execute(
        "SELECT * FROM fn_income_tax_summary_fy(%s,%s)",
        (society_id, fy), fetch_all=True,
    ) or []

    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet(title="Mutuality Summary")
    _apply_header(ws, 2, _COL_WIDTHS)
    _write_mutuality_sheet(ws, rows, fy)

    filename = f"{filename_prefix}_FY{fy}-{fy+1}.xlsx"
    wb._it_filename = filename  # type: ignore[attr-defined]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
