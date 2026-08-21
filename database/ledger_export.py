# database/ledger_export.py
"""
Ledger Excel Generator — EstateHub
====================================
Produces a per-account ledger in the traditional non-paired format,
matching the ld.xlsx reference layout supplied 2026-08:

  Date | A/c | Description | CB F No. | Debit | Credit | Dr or Cr | Balance

DEPENDS ON: fn_account_ledger_fy (estatehub.sql SECTION 5), which already
implements the full year-end cascade — FY-scoped BF resolution, per-
transaction entry_side netting, the depreciation split for depreciable
accounts, and the C/F-to-hierarchy-parent closing row — via row_type
'bf' | 'txn' | 'full_depreciation' | 'half_depreciation' | 'closing'
(2026-08: the single 'depreciation' tag split into these two, matching
that function's own Full-rate/Half-rate-post-1-Sep row split). This
module is a thin Excel-formatting layer over that function's output; it
does NOT reimplement any of the closing arithmetic itself.

Entity/user scoping: NONE. fn_account_ledger_fy takes no entity_id /
entity_role — it's an account-level ledger (every transaction posted to
this one account, across the whole society), not an entity-scoped one.
This matches the live "Account Ledger" drilldown view (loaders.py's
`entity == "ledger"` branch, reached via profile_account -> show_ledger),
which has never taken entity scoping either. Both are consistent with
`ledger` being an admin-only entity (renderers.py's _PORTAL_PERMS gives
every non-admin role an empty permission set for it) — no other portal
can reach a ledger export, so there is nothing to scope by entity in the
first place.

Full-book export (2026-08, added): 'Export Ledger' on the Ledger Index
card now calls generate_ledger_index_excel(), which builds a single
workbook covering the whole chart of accounts for the FY — Index sheet,
one ledger sheet per account (including Dep / InExp / CapAc, which are
just ordinary accounts, id 231 / 23 / 2), and a Bal sheet — matching the
structure of the CB2024-2025.xlsx reference workbook supplied 2026-08.
Both this and generate_ledger_excel share the same per-account sheet
writer (_write_ledger_sheet) so the two exports stay visually identical.

Account hierarchy note: `accounts` is rooted at id=1 (tab_name='Bal',
has_bf=False) — a pure structural node with no transactions of its own,
whose direct children (CapAc, LAT, CurLb, ImAs, MAs, LAG, SDr, SCr) are
the top-level Balance Sheet categories. generate_ledger_index_excel
gives every *other* account its own fn_account_ledger_fy sheet, and
builds the Bal sheet separately from fn_fy_closing_report's depth=1
rows (those same direct children), rather than trying to run
fn_account_ledger_fy against the structural root itself.
"""

from __future__ import annotations
import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

_FONT_BODY   = Font(name="Arial", size=9)
_FONT_HEADER = Font(name="Arial", size=9, bold=True)
_FONT_TITLE  = Font(name="Arial", size=10, bold=True)
_FONT_BFCF   = Font(name="Arial", size=9, bold=True, italic=True)
_FONT_LINK   = Font(name="Arial", size=9, color="FF0563C1", underline="single")
_FONT_TOTAL  = Font(name="Arial", size=9, bold=True)

_FILL_HEADER = PatternFill("solid", fgColor="D9E1F2")
_FILL_BF     = PatternFill("solid", fgColor="E2EFDA")
_FILL_CF     = PatternFill("solid", fgColor="FCE4D6")
_FILL_DEP    = PatternFill("solid", fgColor="FFF2CC")

_ALIGN_C = Alignment(horizontal="center", vertical="center")
_ALIGN_L = Alignment(horizontal="left",   vertical="center")
_ALIGN_R = Alignment(horizontal="right",  vertical="center")

_THIN = Side(style="thin")
_BORDER_ALL = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_FMT_DATE = "DD-MMM-YY"
_FMT_AMT  = '#,##0.00;[Red](#,##0.00);"-"'

_COL_WIDTHS = {"A": 11, "B": 12, "C": 32, "D": 8, "E": 11, "F": 11, "G": 8, "H": 12}

_ROW_STYLE = {
    "bf":                (_FONT_BFCF, _FILL_BF),
    "closing":           (_FONT_BFCF, _FILL_CF),
    "depreciation":      (_FONT_BODY, _FILL_DEP),
    "full_depreciation": (_FONT_BODY, _FILL_DEP),
    "half_depreciation": (_FONT_BODY, _FILL_DEP),
    "txn":               (_FONT_BODY, PatternFill()),
}

# The structural root of the chart of accounts (see module docstring) —
# excluded from the per-account sheet loop in generate_ledger_index_excel
# and handled separately as the Bal sheet.
_ROOT_TAB_NAME = "Bal"

# Closing/rollup accounts that must appear last, in this exact order,
# rather than sorted alphabetically with everything else. Bal (the
# structural root) isn't in this list — it's built and appended
# separately after the main account loop, so it always lands last.
_TAIL_ORDER = ["Dep", "InExp", "CapAc"]


def _account_sort_key(acc: dict):
    tab = (acc.get("tab_name") or acc.get("name") or "").strip()
    if tab in _TAIL_ORDER:
        return (1, _TAIL_ORDER.index(tab), "")
    return (0, 0, tab.lower())


def _unique_sheet_title(base: str, used_titles: set[str]) -> str:
    """
    Excel sheet titles must be <=31 chars and unique within a workbook.
    tab_name is already short and unique per-account in the schema, but
    this guards against any edge case (blank tab_name, collisions after
    truncation) rather than letting openpyxl raise on wb.create_sheet.
    """
    base = (base or "Sheet").strip()[:31] or "Sheet"
    title = base
    n = 2
    while title in used_titles:
        suffix = f"_{n}"
        title = base[: 31 - len(suffix)] + suffix
        n += 1
    used_titles.add(title)
    return title


def _children_rollup_rows(children_closing: list[dict], fy: int) -> list[dict]:
    """
    Converts this account's direct children (rows from
    fn_fy_closing_report where parent_account_id == this account's id)
    into row dicts shaped like fn_account_ledger_fy's own output, so
    _write_ledger_sheet can render them identically.

    fn_account_ledger_fy only reflects transactions posted with
    acc_id = that exact account — it has no visibility into a child
    account's activity. For header/rollup accounts (InExp, CapAc, Bal,
    and any other node with children — e.g. CurAs, BkAc, MAs, IncOther)
    that's not enough on its own: the CB2024-2025.xlsx reference shows
    each child's net contribution as its own line (e.g. InExp lists
    'Dep', 'Salary', 'Misc', ... as separate rows). Those come from
    fn_fy_closing_report instead, which already aggregates each child's
    whole subtree — this just reformats that into ledger rows, using
    the same Cr-positive running-balance convention as the rest of the
    workbook (a credit increases the running balance, a debit reduces
    it — matching estatehub.sql's internal sign convention, not
    necessarily the child's own natural Dr/Cr side).
    """
    fy_end = date(fy + 1, 3, 31)
    rows = []
    running = 0.0
    for c in sorted(children_closing, key=lambda x: x.get("sort_path") or ""):
        side = c.get("display_side")
        amount = float(c.get("display_amount") or 0)
        if amount == 0:
            continue
        running += amount if side == "Cr" else -amount
        rows.append({
            "row_date": fy_end,
            "account_name": c.get("tab_name"),
            "particulars": c.get("account_name"),
            "debit": amount if side == "Dr" else None,
            "credit": amount if side == "Cr" else None,
            "running_balance": running,
            "row_type": "txn",
        })
    return rows


def _closing_transfer_row(own_closing_row: dict, parent_tab_name: str | None, fy: int, running_balance: float) -> dict | None:
    """
    Builds this account's own 'C/F -> {parent}' row from its
    fn_fy_closing_report entry (total_closing/display_side/display_amount,
    which already fold in every child's contribution) — the header-
    account equivalent of the per-transaction C/F row fn_account_ledger_fy
    writes for leaf accounts. Returns None for the ultimate root (no
    parent to close into) or a nil balance (nothing to transfer).
    """
    if not parent_tab_name or not own_closing_row:
        return None
    side = own_closing_row.get("display_side")
    amount = float(own_closing_row.get("display_amount") or 0)
    if amount == 0:
        return None
    return {
        "row_date": date(fy + 1, 3, 31),
        "account_name": own_closing_row.get("tab_name"),
        "particulars": f"C/F -> {parent_tab_name}",
        # A Cr closing balance transfers out as a debit (closing the
        # account to nil) and vice versa — same convention as
        # fn_account_ledger_fy's own closing rows.
        "debit": amount if side == "Cr" else None,
        "credit": amount if side == "Dr" else None,
        "running_balance": 0.0,
        "row_type": "closing",
        "parent_name": parent_tab_name,
    }


def _write_ledger_sheet(ws, acc: dict, ledger_rows: list[dict], fy: int) -> int:
    """
    Writes the standard 8-column ledger layout for one account onto `ws`:

      Date | A/c | Description | LD F No. | Debit | Credit | Dr or Cr | Balance

    Shared by generate_ledger_excel (single-account export) and
    generate_ledger_index_excel (full ledger-book export), so both stay
    visually identical. Returns the next free row (useful for callers
    that append a footer below the ledger, e.g. InExp's Expenditure/
    Receipts/Profit% summary).
    """
    tab = acc.get("tab_name") or acc["name"]
    asst_year = f"{fy}-{fy + 1}"

    for col_letter, width in _COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 6
    title = {
        1: f"{tab}.xlsx", 3: acc["name"], 5: "LEDGER",
        7: "Asst. Yr.", 8: asst_year,
    }
    for col, val in title.items():
        cell = ws.cell(row=2, column=col, value=val)
        cell.font, cell.alignment = _FONT_TITLE, _ALIGN_C
    ws.row_dimensions[3].height = 6

    headers = {1: "Date", 2: "A/c", 3: "Description", 4: "LD F No.",
               5: "Debit", 6: "Credit", 7: "Dr or Cr", 8: "Balance"}
    for col, hdr in headers.items():
        cell = ws.cell(row=4, column=col, value=hdr)
        cell.font, cell.fill = _FONT_HEADER, _FILL_HEADER
        cell.alignment, cell.border = _ALIGN_C, _BORDER_ALL
    ws.row_dimensions[5].height = 6

    current_row = 6
    for r in ledger_rows:
        row_type = r.get("row_type") or "txn"
        font, fill = _ROW_STYLE.get(row_type, (_FONT_BODY, PatternFill()))
        debit  = float(r["debit"])  if r.get("debit")  else None
        credit = float(r["credit"]) if r.get("credit") else None
        balance = float(r.get("running_balance") or 0)
        drcr = None
        if row_type in ("bf", "txn", "depreciation", "full_depreciation", "half_depreciation"):
            drcr = "Dr" if debit and not credit else ("Cr" if credit and not debit else None)

        # parent_name from fn_account_ledger_fy is already tab_name-first
        # (COALESCE(p.tab_name, p.name, '--')) — see estatehub.sql — so no
        # extra resolution needed here for the closing row's A/c column.
        row = {
            1: r.get("row_date"), 2: r.get("parent_name") if row_type == "closing" else (r.get("account_name") or tab),
            3: r.get("particulars") or "",
            5: debit, 6: credit, 7: drcr, 8: abs(balance) if balance is not None else None,
        }
        for col, val in row.items():
            cell = ws.cell(row=current_row, column=col, value=val)
            cell.font, cell.fill = font, fill
            cell.border = _BORDER_ALL
            cell.alignment = _ALIGN_R if col in (5, 6, 8) else _ALIGN_L
            if col == 1 and val:
                cell.number_format = _FMT_DATE
            elif col in (5, 6, 8):
                cell.number_format = _FMT_AMT
        current_row += 1

    ws.freeze_panes = "A6"
    return current_row


def _add_income_expenditure_footer(ws, ledger_rows: list[dict], next_row: int) -> None:
    """
    Appends the 'Expenditure => X    Y <= Receipts' + 'Profit%' summary
    rows the InExp sheet carries in the CB2024-2025.xlsx reference
    (rows 14/16 there). Expenditure = sum of debit on 'txn' rows,
    Receipts = sum of credit on 'txn' rows — B/F and closing rows are
    excluded since they're not the year's actual income/expenditure.
    """
    expenditure = sum(float(r["debit"])  for r in ledger_rows
                       if (r.get("row_type") or "txn") == "txn" and r.get("debit"))
    receipts    = sum(float(r["credit"]) for r in ledger_rows
                       if (r.get("row_type") or "txn") == "txn" and r.get("credit"))

    row = next_row + 1
    cell = ws.cell(row=row, column=4, value="Expenditure =>")
    cell.font, cell.alignment = _FONT_TOTAL, _ALIGN_R
    cell = ws.cell(row=row, column=5, value=expenditure)
    cell.font, cell.alignment, cell.number_format = _FONT_TOTAL, _ALIGN_R, _FMT_AMT
    cell = ws.cell(row=row, column=6, value=receipts)
    cell.font, cell.alignment, cell.number_format = _FONT_TOTAL, _ALIGN_R, _FMT_AMT
    cell = ws.cell(row=row, column=7, value="<= Receipts")
    cell.font, cell.alignment = _FONT_TOTAL, _ALIGN_L

    if receipts:
        profit_pct = (receipts - expenditure) / receipts
        row2 = row + 2
        cell = ws.cell(row=row2, column=4, value="Profit%")
        cell.font, cell.alignment = _FONT_TOTAL, _ALIGN_R
        cell = ws.cell(row=row2, column=5, value=profit_pct)
        cell.font, cell.alignment, cell.number_format = _FONT_TOTAL, _ALIGN_R, "0.00%"


def generate_ledger_excel(
    db,
    society_id: int,
    fy: int,
    account_id: int,
) -> bytes:
    """
    Builds a single-account ledger sheet for the given FY, sourced
    entirely from fn_account_ledger_fy(society_id, account_id, fy) — see
    module docstring for why no entity_id/entity_role params are taken.
    """
    from database.db_manager import db as _db
    if db is None:
        db = _db

    acc = db._execute(
        "SELECT id, name, tab_name FROM accounts WHERE id=%s AND society_id=%s",
        (account_id, society_id), fetch_one=True,
    )
    if not acc:
        raise ValueError(f"Account {account_id} not found for society {society_id}")

    tab = acc.get("tab_name") or acc["name"]

    # Insertion order out of fn_account_ledger_fy is already correct
    # (bf -> txns by date -> depreciation -> closing); no ORDER BY here,
    # since sorting by row_date alone would shuffle the depreciation and
    # closing rows, which share the same FY-end date.
    ledger_rows = db._execute(
        "SELECT * FROM fn_account_ledger_fy(%s,%s,%s)",
        (society_id, account_id, fy), fetch_all=True,
    ) or []

    wb = Workbook()
    ws = wb.active
    ws.title = tab[:31]

    next_row = _write_ledger_sheet(ws, acc, ledger_rows, fy)
    if (acc.get("tab_name") or "").lower() == "inexp":
        _add_income_expenditure_footer(ws, ledger_rows, next_row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def generate_ledger_index_excel(
    db,
    society_id: int,
    fy: int,
) -> bytes:
    """
    Builds the full ledger book for the FY — one workbook covering the
    whole chart of accounts, matching the CB2024-2025.xlsx reference
    layout: an Index sheet (S No. | Account | Page No., each row linking
    to its sheet), a ledger sheet per account (via fn_account_ledger_fy —
    this is what carries Dep / InExp / CapAc, which are just ordinary
    accounts, ids 231 / 23 / 2 in the seed chart of accounts), and a
    closing Bal sheet built from fn_fy_closing_report's depth=1 rows
    (the root account's direct children — CapAc, LAT, CurLb, ImAs, MAs,
    LAG, SDr, SCr — the standard top-level Balance Sheet categories).

    This is what 'Export Ledger' on the Ledger Index card calls.
    """
    from database.db_manager import db as _db
    if db is None:
        db = _db

    # fn_accounts_hierarchy orders by tree.sort_path internally (parent
    # immediately followed by its own children) — no ORDER BY needed here.
    accounts = db._execute(
        "SELECT * FROM fn_accounts_hierarchy(%s)",
        (society_id,), fetch_all=True,
    ) or []

    root = next((a for a in accounts if not a.get("parent_account_id")), None)
    non_root_accounts = [a for a in accounts
                          if (a.get("tab_name") or "").strip().lower() != _ROOT_TAB_NAME.lower()]

    # Sheet order: alphabetical by tab_name, but ending with the four
    # closing/rollup accounts in a fixed sequence — Dep, InExp, CapAc,
    # then Bal (Bal is the structural root, appended separately below
    # the loop, so it naturally lands last regardless of this sort).
    non_root_accounts = sorted(non_root_accounts, key=_account_sort_key)

    closing_rows = db._execute(
        "SELECT * FROM fn_fy_closing_report(%s,%s) ORDER BY sort_path",
        (society_id, fy), fetch_all=True,
    ) or []
    closing_by_id = {r["account_id"]: r for r in closing_rows}
    children_by_parent: dict[int, list[dict]] = {}
    for r in closing_rows:
        pid = r.get("parent_account_id")
        if pid is not None:
            children_by_parent.setdefault(pid, []).append(r)

    wb = Workbook()
    used_titles: set[str] = set()

    index_ws = wb.active
    index_ws.title = _unique_sheet_title("Index", used_titles)

    entries: list[tuple[str, str]] = []  # (sheet_title, display_name), in page order

    def _build_account_rows(acc_id: int, own_ledger_rows: list[dict], parent_tab_name: str | None) -> list[dict]:
        """
        Own fn_account_ledger_fy rows (bf/txn/depreciation — never a
        'closing' row here, since fn_account_ledger_fy only emits one
        when the account has its own direct transactions) + this
        account's children rolled up from fn_fy_closing_report (empty
        for true leaf accounts) + a single synthetic closing row
        transferring the account's own total_closing to its parent.
        """
        rows = [r for r in own_ledger_rows if (r.get("row_type") or "txn") != "closing"]
        own_closing_row_from_ledger_fy = next(
            (r for r in own_ledger_rows if (r.get("row_type") or "") == "closing"), None
        )
        children = children_by_parent.get(acc_id, [])
        if children:
            rows.extend(_children_rollup_rows(children, fy))
        if children and parent_tab_name:
            # Header/rollup account: derive the C/F row from its own
            # aggregated fn_fy_closing_report entry (folds in every
            # child), rather than fn_account_ledger_fy's row (which
            # only sees direct postings and would be nil here).
            transfer = _closing_transfer_row(closing_by_id.get(acc_id), parent_tab_name, fy, 0.0)
            if transfer:
                rows.append(transfer)
        elif own_closing_row_from_ledger_fy:
            # Leaf account: fn_account_ledger_fy already computed its
            # correct C/F row from its own direct transactions.
            rows.append(own_closing_row_from_ledger_fy)
        return rows

    # ── one ledger sheet per account ────────────────────────────────────
    for acc in non_root_accounts:
        sheet_title = _unique_sheet_title(acc.get("tab_name") or acc["name"], used_titles)
        ws = wb.create_sheet(sheet_title)

        own_ledger_rows = db._execute(
            "SELECT * FROM fn_account_ledger_fy(%s,%s,%s)",
            (society_id, acc["id"], fy), fetch_all=True,
        ) or []
        combined_rows = _build_account_rows(acc["id"], own_ledger_rows, acc.get("parent_tab_name"))

        next_row = _write_ledger_sheet(ws, acc, combined_rows, fy)
        if (acc.get("tab_name") or "").lower() == "inexp":
            _add_income_expenditure_footer(ws, combined_rows, next_row)

        entries.append((sheet_title, acc["name"]))

    # ── Bal sheet (root's own direct children, same rollup logic) ──────
    bal_title = _unique_sheet_title(_ROOT_TAB_NAME, used_titles)
    bal_ws = wb.create_sheet(bal_title)
    root_rows = _children_rollup_rows(children_by_parent.get((root or {}).get("id"), []), fy)
    _write_ledger_sheet(bal_ws, {"name": (root or {}).get("name") or "Balance Sheet",
                                  "tab_name": _ROOT_TAB_NAME}, root_rows, fy)
    entries.append((bal_title, (root or {}).get("name") or "Balance Sheet"))

    # ── Index sheet ──────────────────────────────────────────────────────
    _write_index_sheet(index_ws, entries)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _write_index_sheet(ws, entries: list[tuple[str, str]]) -> None:
    for col_letter, width in {"A": 8, "B": 40, "C": 10}.items():
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 6
    cell = ws.cell(row=2, column=2, value="Index")
    cell.font, cell.alignment = _FONT_TITLE, _ALIGN_C
    ws.row_dimensions[3].height = 6

    headers = {1: "S No.", 2: "Account", 3: "Page No."}
    for col, hdr in headers.items():
        cell = ws.cell(row=4, column=col, value=hdr)
        cell.font, cell.fill = _FONT_HEADER, _FILL_HEADER
        cell.alignment, cell.border = _ALIGN_C, _BORDER_ALL
    ws.row_dimensions[5].height = 6

    for i, (sheet_title, display_name) in enumerate(entries, start=1):
        page_no = i
        row = 5 + i
        c1 = ws.cell(row=row, column=1, value=i)
        c1.alignment, c1.border = _ALIGN_C, _BORDER_ALL
        c1.font = _FONT_BODY

        c2 = ws.cell(row=row, column=2, value=display_name)
        c2.alignment, c2.border = _ALIGN_L, _BORDER_ALL
        c2.font = _FONT_LINK
        c2.hyperlink = f"#'{sheet_title}'!A1"

        c3 = ws.cell(row=row, column=3, value=page_no)
        c3.alignment, c3.border = _ALIGN_C, _BORDER_ALL
        c3.font = _FONT_BODY


def generate_fy_closing_excel(
    db,
    society_id: int,
    fy: int,
) -> bytes:
    """
    Builds a single-sheet FY Closing Report workbook sourced from
    fn_fy_closing_report(society_id, fy). Superseded for the Ledger
    Index's 'Export Ledger' button by generate_ledger_index_excel()
    (2026-08), which produces the full multi-sheet ledger book instead
    of just this summary; kept standalone since it's a useful flat
    one-sheet view in its own right.
    """
    from database.db_manager import db as _db
    if db is None:
        db = _db

    rows = db._execute(
        "SELECT * FROM fn_fy_closing_report(%s,%s) ORDER BY sort_path",
        (society_id, fy), fetch_all=True,
    ) or []

    wb = Workbook()
    ws = wb.active
    ws.title = "FY Closing"

    for col_letter, width in {"A": 28, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14, "G": 8}.items():
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 6
    title = {1: "FY Closing Report", 3: f"FY {fy}-{str(fy + 1)[-2:]}", 5: "Asst. Yr."}
    for col, val in title.items():
        cell = ws.cell(row=2, column=col, value=val)
        cell.font, cell.alignment = _FONT_TITLE, _ALIGN_C
    ws.row_dimensions[3].height = 6

    headers = {1: "Account", 2: "B/F", 3: "Movement", 4: "Dep.", 5: "Own Closing",
               6: "Total Closing", 7: "Dr/Cr"}
    for col, hdr in headers.items():
        cell = ws.cell(row=4, column=col, value=hdr)
        cell.font, cell.fill = _FONT_HEADER, _FILL_HEADER
        cell.alignment, cell.border = _ALIGN_C, _BORDER_ALL
    ws.row_dimensions[5].height = 6

    current_row = 6
    for r in rows:
        drcr = r.get("display_side")
        row = {
            1: r.get("account_name"),
            2: float(r.get("own_bf") or 0),
            3: float(r.get("own_movement") or 0),
            4: float(r.get("depreciation_charge") or 0),
            5: float(r.get("own_closing") or 0),
            6: float(r.get("display_amount") or 0),
            7: drcr,
        }
        for col, val in row.items():
            cell = ws.cell(row=current_row, column=col, value=val)
            cell.font, cell.fill = _FONT_BODY, PatternFill()
            cell.border = _BORDER_ALL
            cell.alignment = _ALIGN_R if col in (2, 3, 4, 5, 6) else _ALIGN_L
            if col in (2, 3, 4, 5, 6):
                cell.number_format = _FMT_AMT
            if col == 7 and val:
                cell.font = Font(name="Arial", size=9, bold=True,
                                 color="FFFF0000" if val == "Dr" else "FF008000")
        current_row += 1

    ws.freeze_panes = "A6"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
